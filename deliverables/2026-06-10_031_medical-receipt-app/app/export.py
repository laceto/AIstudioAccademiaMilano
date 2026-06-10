"""Export receipts to Excel (.xlsx) and PDF summary."""
from __future__ import annotations

import io
from datetime import date
from typing import Sequence

from app.constants import DEDUCTION_RATE, FRANCHISE_EUR
from app.models import Receipt
from app.schemas import EXPENSE_TYPES


def to_excel(receipts: Sequence[Receipt], fiscal_year: int) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("openpyxl not installed") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Spese Sanitarie {fiscal_year}"

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(color="FFFFFF", bold=True)
    headers = [
        "Data", "N. Scontrino", "Fornitore", "P.IVA/CF",
        "Tipo Spesa", "Descrizione", "Metodo Pagamento",
        "Importo (€)", "Detraibile (€)", "Detraibile 730", "Note",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not receipts:
        ws.cell(row=2, column=1, value="Nessuna ricevuta confermata per questo anno fiscale.")
    else:
        for row_idx, r in enumerate(receipts, 2):
            ded = float(r.deductible_amount or r.total_amount or 0) if r.tax_deductible else 0.0
            ws.cell(row=row_idx, column=1, value=r.date or "")
            ws.cell(row=row_idx, column=2, value=r.receipt_number or "")
            ws.cell(row=row_idx, column=3, value=r.provider_name)
            ws.cell(row=row_idx, column=4, value=r.provider_tax_id or "")
            ws.cell(row=row_idx, column=5, value=EXPENSE_TYPES.get(r.expense_type, r.expense_type))
            ws.cell(row=row_idx, column=6, value=r.description or "")
            ws.cell(row=row_idx, column=7, value=r.payment_method or "")
            ws.cell(row=row_idx, column=8, value=float(r.total_amount or 0))
            ws.cell(row=row_idx, column=9, value=ded)
            ded_label = "Sì" if r.tax_deductible is True else ("Da verificare" if r.tax_deductible is None else "No")
            ws.cell(row=row_idx, column=10, value=ded_label)
            ws.cell(row=row_idx, column=11, value=r.notes or "")

        last_data_row = len(receipts) + 1
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=7, value="TOTALE").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{last_data_row})").font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{last_data_row})").font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    summary_ws = wb.create_sheet("Riepilogo 730")
    total_deductible = sum(
        float(r.deductible_amount or r.total_amount or 0)
        for r in receipts if r.tax_deductible is True
    )
    taxable_base = max(0, total_deductible - FRANCHISE_EUR)
    estimated_saving = round(taxable_base * DEDUCTION_RATE, 2)
    unknown_count = sum(1 for r in receipts if r.tax_deductible is None)

    summary_data = [
        ("Anno fiscale", fiscal_year),
        ("Numero ricevute confermate", len(receipts)),
        ("Totale spese (€)", round(sum(float(r.total_amount or 0) for r in receipts), 2)),
        ("Totale detraibile lordo (€)", round(total_deductible, 2)),
        ("Franchigia 730 (€)", FRANCHISE_EUR),
        ("Base imponibile detraibile (€)", round(taxable_base, 2)),
        ("Detrazione stimata 19% (€)", estimated_saving),
        ("", ""),
        ("Ricevute con detraibilità da verificare", unknown_count),
        ("", ""),
        ("Nota", "La detrazione effettiva dipende dal reddito e dalla situazione fiscale individuale. Consulta il tuo CAF."),
    ]
    for i, (label, value) in enumerate(summary_data, 1):
        ws_cell = summary_ws.cell(row=i, column=1, value=label)
        ws_cell.font = Font(bold=True)
        summary_ws.cell(row=i, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf_summary(receipts: Sequence[Receipt], fiscal_year: int) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:
        raise RuntimeError("reportlab not installed") from e

    total_amount = sum(float(r.total_amount or 0) for r in receipts)
    total_deductible = sum(
        float(r.deductible_amount or r.total_amount or 0)
        for r in receipts if r.tax_deductible is True
    )
    taxable_base = max(0.0, total_deductible - FRANCHISE_EUR)
    estimated_saving = round(taxable_base * DEDUCTION_RATE, 2)
    unknown_count = sum(1 for r in receipts if r.tax_deductible is None)

    by_type: dict[str, float] = {}
    for r in receipts:
        label = EXPENSE_TYPES.get(r.expense_type, r.expense_type)
        by_type[label] = by_type.get(label, 0.0) + float(r.total_amount or 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#1A73E8")
    story = []

    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=brand, fontSize=20, spaceAfter=4)
    story.append(Paragraph(f"Spese Sanitarie {fiscal_year}", title_style))
    story.append(Paragraph(f"Generato il {date.today().strftime('%d/%m/%Y')}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    if not receipts:
        story.append(Paragraph("Nessuna ricevuta confermata per questo anno fiscale.", styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    summary_data = [
        ["Descrizione", "Valore"],
        ["Numero ricevute", str(len(receipts))],
        ["Totale spese sanitarie", f"€ {total_amount:,.2f}"],
        ["Totale detraibile lordo", f"€ {total_deductible:,.2f}"],
        ["Franchigia annua 730", f"€ {FRANCHISE_EUR:,.2f}"],
        ["Base imponibile detraibile", f"€ {taxable_base:,.2f}"],
        ["Detrazione stimata (19%)", f"€ {estimated_saving:,.2f}"],
    ]
    if unknown_count:
        summary_data.append([f"⚠ Detraibilità da verificare", f"{unknown_count} ricevute"])

    summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5*cm))

    if by_type:
        cat_data = [["Categoria", "Importo (€)"]] + [
            [cat, f"€ {amt:,.2f}"]
            for cat, amt in sorted(by_type.items(), key=lambda x: -x[1])
        ]
        cat_table = Table(cat_data, colWidths=[10*cm, 6*cm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.5*cm))

    receipt_data = [["Data", "Fornitore", "Tipo", "Importo", "Detraibile", "Stato"]]
    for r in receipts:
        ded = float(r.deductible_amount or r.total_amount or 0) if r.tax_deductible is True else 0.0
        ded_label = f"€{ded:.2f}" if r.tax_deductible is True else ("?" if r.tax_deductible is None else "No")
        receipt_data.append([
            r.date or "—",
            (r.provider_name[:30] + "…") if len(r.provider_name) > 30 else r.provider_name,
            EXPENSE_TYPES.get(r.expense_type, r.expense_type)[:15],
            f"€{float(r.total_amount or 0):.2f}",
            ded_label,
            "Confermato" if r.status == "confirmed" else "In revisione",
        ])
    col_w = [2.2*cm, 6.5*cm, 3*cm, 2.3*cm, 2.3*cm, 2.5*cm]
    receipt_table = Table(receipt_data, colWidths=col_w, repeatRows=1)
    receipt_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("ALIGN", (3, 0), (4, -1), "RIGHT"),
    ]))
    story.append(receipt_table)
    story.append(Spacer(1, 0.5*cm))

    disclaimer = ParagraphStyle("disc", parent=styles["Normal"], fontSize=7, textColor=colors.grey)
    story.append(Paragraph(
        "Nota: I valori indicati sono stime. La detrazione effettiva dipende dalla situazione fiscale individuale. "
        "Consulta il tuo CAF o commercialista per la dichiarazione dei redditi. "
        "Le ricevute con '?' nella colonna Detraibile richiedono verifica manuale prima della presentazione del 730.",
        disclaimer,
    ))

    doc.build(story)
    return buf.getvalue()
